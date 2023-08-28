#!/usr/bin/env python 
# -*- coding: utf-8 -*-
'''
Time    : 2023/8/22 16:12
Author  : zhuchunjin
Email   : chunjin.zhu@taurentech.net
File    : load_ui_panel.py
Software: PyCharm
'''
import os.path
import re
import sys
import time
from binascii import hexlify
from ctypes import byref, c_ubyte

import numpy as np
import win32com
from win32com.client import Dispatch

import xlrd
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtWidgets import QFileDialog, QInputDialog, QMessageBox

from script import ControlSPI
from script.ControlSPI import VSI_CloseDevice
from script.mem_read import MainDialog1
from script.xihe_UI import Ui_MainWindow
from image import status_rc
from image import structure_rc
import openpyxl as op


# 生成资源文件目录访问路径
def resource_path(relative_path):
    if getattr(sys, 'frozen', False):  # 是否Bundle Resource
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class load_xihepango_panel(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.dac_file_path = None
        self.dac_pll_config_file_path = None
        self.choose_Electric = None
        self.amp_dac = None
        self.dobule_space = None
        self.dac_send_freq = None
        self.single_two = None
        self.dac_mode = None
        self.dac_chn = None
        self.dac_setting_tile = 0
        self.dac_ref = None
        self.dac_sample_rate = None
        self.dac_pll_config_tile = None
        self.choose_path = "ADC"
        self.setupUi(self)
        self.dac_sample_lineEdit.setText("10000")
        self.dac_ref_lineEdit.setText("100")
        self.dac_freq_lineEdit.setText("3500")
        self.amp_lineEdit.setText("0")
        self.two_space_lineEdit.setText("20")
        self.adc_sample_pll_lineEdit.setText("3000")
        self.adc_ref_lineEdit.setText("100")
        self.dac_pll_config_groupBox.setEnabled(False)
        self.dac_config_groupBox.setEnabled(False)
        self.dac_send_groupBox.setEnabled(False)
        self.two_space_lineEdit.setEnabled(False)
        self.signal_choose_comboBox.currentIndexChanged.connect(self.choose_chnannel_setting)
        self.write_button.clicked.connect(self.write_addr)
        self.read_button.clicked.connect(self.read_addr)
        self.dac_pll_config_lineEdit.setVisible(False)
        self.dac_file_lineEdit.setVisible(False)
        self.adc_pll_file_lineEdit.setVisible(False)
        self.adc_file_lineEdit.setVisible(False)
        self.dac_pll_toolButton.setVisible(False)
        self.adc_pll_config_toolButton.setVisible(False)
        self.adc_file_toolButton.setVisible(False)
        self.dac_file_toolButton.setVisible(False)
        self.dac_pll_config_lineEdit.setText(resource_path(os.path.join("source", "pll_pango_dac_v0.2.xlsx")))
        self.dac_file_lineEdit.setText(resource_path(os.path.join("source", "Pango-DAC_v0.2.xlsx")))
        self.adc_pll_file_lineEdit.setText(resource_path(os.path.join("source", "pll-pango-adc.xlsx")))
        self.adc_file_lineEdit.setText(resource_path(os.path.join("source", "xihev300_adc_reg_001.xlsx")))
        self.adc_dac_comboBox.currentIndexChanged.connect(self.choose_adc_or_dac)
        self.clear_log_btn.clicked.connect(self.clear_log_content)
        self.view_all_log_btn.clicked.connect(self.get_log_content)
        self.connect_btn.clicked.connect(self.init_spi_config)
        self.release_btn.clicked.connect(self.spi_release)
        self.load_sheet_btn.clicked.connect(self.load_test_seq)
        self.dac_pll_tile_comboBox.currentIndexChanged.connect(self.dac_tile_change_ton)
        self.dac_tile_comboBox.currentIndexChanged.connect(self.dac_tile_change_ton)
        self.adc_tile_comboBox.currentIndexChanged.connect(self.adc_tile_change_ton)
        self.adc_pll_tile_comboBox.currentIndexChanged.connect(self.adc_tile_change_ton)
        self.update_btn.clicked.connect(self.excel_load_change)
        # self.adc_pll_config_toolButton.clicked.connect(lambda: self.file_path_load(1))
        # self.dac_pll_toolButton.clicked.connect(lambda: self.file_path_load(2))
        # self.adc_file_toolButton.clicked.connect(lambda: self.file_path_load(3))
        # self.dac_file_toolButton.clicked.connect(lambda: self.file_path_load(4))
        self.adc_mem_read_btn.clicked.connect(self.wirte_to_txt_sample)
        self.base_addr_dict = {"DAC": {0: 0x20000, 1: 0x60000, 2: 0xa0000, 3: 0xe0000}, "ADC": {0: 0x00000, 1: 0x40000,
                                                                                                2: 0x80000, 3: 0xc0000}}

    def dac_tile_change_ton(self, index):
        self.dac_pll_tile_comboBox.setCurrentIndex(index)
        self.dac_tile_comboBox.setCurrentIndex(index)

    def adc_tile_change_ton(self, index):
        self.adc_pll_tile_comboBox.setCurrentIndex(index)
        self.adc_tile_comboBox.setCurrentIndex(index)

    def file_path_load(self, id):
        test_seq_file, filetype = QFileDialog.getOpenFileName(self, "choose file", "./",
                                                              "excel Files (*.xlsx);;All Files (*)")  # 设置文件扩展名过滤,注意用双分号间隔
        if test_seq_file == "":
            return
        else:
            if id == 1:
                self.adc_pll_file_lineEdit.setText(test_seq_file)
            elif id == 2:
                self.dac_pll_config_lineEdit.setText(test_seq_file)
            elif id == 3:
                self.adc_file_lineEdit.setText(test_seq_file)
            elif id == 4:
                self.dac_file_lineEdit.setText(test_seq_file)

    def excel_load_change(self):
        # try:
        self.textBrowser_normal_log("寄存器配置开始!")
        if self.choose_path == "DAC":
            base_addr = self.base_addr_dict[self.choose_path][self.dac_tile_comboBox.currentIndex()]
            self.read_atom(base_addr + 0x20351)
            self.write_atom(base_addr + 0x20351, 0x0)
            self.read_atom(base_addr + 0x20351)
            # spi.read_atom(0x40200)
            time.sleep(0.1)
            if self.dac_pll_config_lineEdit.text() == "" or self.dac_file_lineEdit.text() == "":
                QMessageBox.information(self, "info", "请先将文件路径确认！")
                return
            self.dac_pll_config_tile = self.dac_pll_tile_comboBox.currentIndex()
            self.dac_sample_rate = int(self.dac_sample_lineEdit.text())
            self.dac_ref = int(self.dac_ref_lineEdit.text())
            self.dac_setting_tile = self.dac_tile_comboBox.currentIndex()
            self.dac_chn = self.dac_chn_comboBox.currentIndex()
            self.dac_mode = self.mode_choose_comboBox.currentIndex()
            self.single_two = self.signal_choose_comboBox.currentIndex()
            self.dac_send_freq = int(self.dac_freq_lineEdit.text())
            self.dobule_space = int(self.two_space_lineEdit.text())
            self.amp_dac = int(self.amp_lineEdit.text())
            self.choose_Electric = self.comboBox.currentIndex()
            ############# M=1
            self.dac_pll_config_file_path = self.dac_pll_config_lineEdit.text()
            self.dac_file_path = self.dac_file_lineEdit.text()
            if not self.checkBox.checkState():
                dac_pll_config_file = op.load_workbook(self.dac_pll_config_file_path)
                table_pll_dac = dac_pll_config_file['寄存器表单']
                if 1000 <= self.dac_sample_rate < 2500:
                    m_value = 4
                elif 2500 <= self.dac_sample_rate < 5000:
                    m_value = 2
                elif 5000 <= self.dac_sample_rate <= 10000:
                    m_value = 1
                else:
                    self.textBrowser_error_log("采样频率不对")
                    return
                table_pll_dac.cell(45, 10, m_value)
                n_value = int(self.dac_sample_rate / self.dac_ref*m_value)
                table_pll_dac.cell(44, 10, n_value)
                if self.dac_ref*n_value > 7500:
                    table_pll_dac.cell(19, 10, 1)
                    table_pll_dac.cell(20, 10, 0)
                else:
                    table_pll_dac.cell(19, 10, 0)
                    table_pll_dac.cell(20, 10, 1)
                dac_pll_config_file.save(self.dac_pll_config_file_path)
                dac_pll_config_file.close()
                self.just_open(self.dac_pll_config_file_path)
            dac_setting_file = op.load_workbook(self.dac_file_path)
            table_dac_setting = dac_setting_file['寄存器配置']
            table_dac_setting.cell(27, 6, self.dac_chn)
            table_dac_setting.cell(60, 7, 1 - self.dac_mode)
            table_dac_setting.cell(37, 8, self.single_two)
            table_dac_setting.cell(30, 7, self.dac_sample_rate / 1000)
            table_dac_setting.cell(3, 4, 1 - self.choose_Electric)
            if not self.single_two:
                table_dac_setting.cell(30, 8, self.dac_send_freq / 1000)
            else:
                table_dac_setting.cell(30, 8, self.dac_send_freq / 1000)
                table_dac_setting.cell(30, 9, self.dac_send_freq / 1000 + self.dobule_space / 1000)
            table_dac_setting.cell(33, 7, self.amp_dac)
            dac_setting_file.save(self.dac_file_path)
            dac_setting_file.close()
            self.just_open(self.dac_file_path)
            self.parser_no_choose_seq_file(self.dac_pll_config_file_path, "寄存器表单", base_addr)
            self.parser_no_choose_seq_file(self.dac_file_path, "寄存器配置", base_addr)
            self.textBrowser_normal_log("更新表格数据成功并加载表格进行读写寄存器！")
        else:
            if self.adc_pll_file_lineEdit.text() == "" or self.adc_file_lineEdit.text() == "":
                QMessageBox.information(self, "info", "请先将文件路径确认！")
                return
            else:
                base_addr = self.base_addr_dict[self.choose_path][self.adc_pll_tile_comboBox.currentIndex()]
                self.read_atom(base_addr + 0x351)
                self.write_atom(base_addr + 0x351, 0x0)
                self.read_atom(base_addr + 0x351)
                self.read_atom(base_addr + 0x40200)
                time.sleep(0.1)
                self.adc_pll_file = self.adc_pll_file_lineEdit.text()
                self.adc_pll_tile = self.adc_pll_tile_comboBox.currentIndex()
                self.adc_sample_rate = int(self.adc_sample_pll_lineEdit.text())
                self.adc_ref = int(self.adc_ref_lineEdit.text())
                self.adc_setting_file = self.adc_file_lineEdit.text()
                self.adc_setting_tile = self.adc_tile_comboBox.currentIndex()
                self.adc_chn_choose = self.adc_chn_choose_comboBox.currentIndex()
                self.adc_speed_choose = self.adc_speed_choose_comboBox.currentIndex()
                ##########
                if not self.checkBox.checkState():
                    adc_pll_config_file = op.load_workbook(self.adc_pll_file)
                    table_pll_adc = adc_pll_config_file['寄存器表单']
                    if 1000 <= self.adc_sample_rate < 2500:
                        m_value = 4
                    elif 2500 <= self.adc_sample_rate < 5000:
                        m_value = 2
                    elif 5000 <= self.adc_sample_rate <= 10000:
                        m_value = 1
                    else:
                        self.textBrowser_error_log("采样频率不对")
                        return
                    table_pll_adc.cell(45, 10, m_value)
                    n_value = int((self.adc_sample_rate / self.adc_ref) * m_value)
                    table_pll_adc.cell(44, 10, n_value)
                    if self.adc_ref*n_value > 7500:
                        table_pll_adc.cell(19, 10, 1)
                        table_pll_adc.cell(20, 10, 0)
                    else:
                        table_pll_adc.cell(19, 10, 0)
                        table_pll_adc.cell(20, 10, 1)
                    adc_pll_config_file.save(self.adc_pll_file)
                    adc_pll_config_file.close()
                    self.just_open(self.adc_pll_file)
                    self.parser_no_choose_seq_file(self.adc_pll_file, "寄存器表单", base_addr)
                ############
                adc_setting_file = op.load_workbook(self.adc_setting_file)

                if self.adc_speed_choose:
                    if self.adc_sample_rate == 6000:
                        table_speed_adc = adc_setting_file.get_sheet_by_name("adc_6g")
                    else:
                        table_speed_adc = adc_setting_file.get_sheet_by_name("adc_6g_dth")
                else:
                    if self.adc_sample_rate == 3000:
                        table_speed_adc = adc_setting_file.get_sheet_by_name("adc_3g")
                    else:
                        table_speed_adc = adc_setting_file.get_sheet_by_name("adc_3g_dth")
                table_speed_adc_rows_num = table_speed_adc.max_row
                for x in range(1, table_speed_adc_rows_num + 1):
                    if table_speed_adc.cell(x, 1).value is not None:
                        temp_addr = int(table_speed_adc.cell(x, 1).value, 16) + self.base_addr_dict[self.choose_path][
                            self.adc_pll_tile_comboBox.currentIndex()]
                        temp_data = int(table_speed_adc.cell(x, 2).value, 16)
                        self.write_atom(temp_addr, temp_data)
                        time.sleep(0.01)
                table_calib_adc = adc_setting_file.get_sheet_by_name("calib")
                table_calib_adc_rows_num = table_calib_adc.max_row

                for x in range(1, table_calib_adc_rows_num + 1):
                    if table_calib_adc.cell(x, self.adc_chn_choose * 2 + 1).value is not None:
                        temp_addr = int(table_calib_adc.cell(x, self.adc_chn_choose * 2 + 1).value, 16) + \
                                    self.base_addr_dict[self.choose_path][self.adc_pll_tile_comboBox.currentIndex()]
                        temp_data = int(table_calib_adc.cell(x, self.adc_chn_choose * 2 + 2).value, 16)
                        self.write_atom(temp_addr, temp_data)
                        time.sleep(0.01)
                adc_setting_file.close()
        time.sleep(5)
        self.textBrowser_normal_log("寄存器配置finish!")
        # except Exception as e:
        #     self.textBrowser_error_log("%s" % e)

    def choose_chnannel_setting(self):
        if self.signal_choose_comboBox.currentIndex() == 0:
            self.two_space_lineEdit.setEnabled(False)
        else:
            self.two_space_lineEdit.setEnabled(True)

    def choose_adc_or_dac(self):
        self.choose_path = self.adc_dac_comboBox.currentText()
        if self.choose_path == "ADC":
            self.dac_pll_config_groupBox.setEnabled(False)
            self.dac_config_groupBox.setEnabled(False)
            self.dac_send_groupBox.setEnabled(False)
            ######3
            self.adc_groupBox.setEnabled(True)
            self.adc_pll_config_groupBox.setEnabled(True)
        else:
            self.adc_groupBox.setEnabled(False)
            self.adc_pll_config_groupBox.setEnabled(False)
            ############
            self.dac_pll_config_groupBox.setEnabled(True)
            self.dac_config_groupBox.setEnabled(True)
            self.dac_send_groupBox.setEnabled(True)

    """
           spi_init 初始化
    """

    def init_spi_config(self):
        # self.spi_release()
        ##########################box#######################################333
        nRet = ControlSPI.VSI_ScanDevice(1)
        # Initialize device
        SPI_Init = ControlSPI.VSI_INIT_CONFIG()
        SPI_Init.ClockSpeed = int(float(self.clk_lineEdit.text()) * 10e6)
        SPI_Init.ControlMode = 3
        SPI_Init.CPHA = 0
        SPI_Init.CPOL = 0
        SPI_Init.LSBFirst = 0
        SPI_Init.MasterMode = 1
        SPI_Init.SelPolarity = 0
        SPI_Init.TranBits = 8
        nRet = ControlSPI.VSI_InitSPI(ControlSPI.VSI_USBSPI, 0, byref(SPI_Init))
        #   Open device
        nRet = ControlSPI.VSI_OpenDevice(ControlSPI.VSI_USBSPI, 0, 0)
        if nRet != ControlSPI.ERR_SUCCESS:
            self.textBrowser_error_log("Open device error!")
            self.condition_label.setPixmap(QtGui.QPixmap(":/images/unlink_status.png"))
        else:
            self.textBrowser_normal_log("Open device success!")
            self.condition_label.setPixmap(QtGui.QPixmap(":/images/link_status.png"))
            # Initialize device
        nRet = ControlSPI.VSI_InitSPI(ControlSPI.VSI_USBSPI, 0, byref(SPI_Init))
        if nRet != ControlSPI.ERR_SUCCESS:
            self.textBrowser_error_log("Initialization device error!")
            self.condition_label.setPixmap(QtGui.QPixmap(":/images/unlink_status.png"))
        else:
            self.textBrowser_normal_log("Initialization device success!")
            self.condition_label.setPixmap(QtGui.QPixmap(":/images/link_status.png"))

    def spi_release(self):
        try:
            VSI_CloseDevice(ControlSPI.VSI_USBSPI, 0)
            self.textBrowser_normal_log("release success")
            self.condition_label.setPixmap(QtGui.QPixmap(":/images/unlink_status.png"))
        except Exception as e:
            self.textBrowser_error_log('%s' % e)

    def textBrowser_normal_log(self, info):
        self.log_textBrowser.append("<font color='black'>" + "{0} {1}".format(time.strftime("%F %T"), info))

    def textBrowser_error_log(self, info):
        self.log_textBrowser.append("<font color='red'>" + '{0} {1}'.format(time.strftime("%F %T"), info))

    def clear_log_content(self):
        self.log_textBrowser.clear()

    def get_log_content(self):
        file_name, file_type = QFileDialog.getSaveFileName(self, "文件保存", "./", "text file (*.txt)")
        if file_name.strip(" ") != "":
            with open(file_name, 'w') as fileOpen:
                fileOpen.write(self.log_textBrowser.toPlainText())

    def read_addr(self):
        now_addr = self.addr_textEdit.text()
        try:
            if re.match('^0x', now_addr):
                addr_read = int(now_addr, 16)
            else:
                addr_read = int(now_addr)
            read_value = self.read_atom(addr_read)
            self.textEdit.setText(
                hex(read_value[0] * 16777216 + read_value[1] * 65536 + read_value[2] * 256 + read_value[3]))
            self.textBrowser_normal_log("read done!")
        except Exception as e:
            self.textBrowser_error_log('read exists err:%s' % e)

    def write_addr(self):
        now_addr = self.addr_textEdit.text()
        try:
            if re.match('^0x', now_addr):
                addr_write = int(now_addr, 16)
            else:
                addr_write = int(now_addr)
            now_value = self.textEdit.text()
            if re.match('^0x', now_value):
                write_value = int(now_value, 16)
            else:
                write_value = int(now_value)
            self.write_atom(addr_write, write_value)
            self.textBrowser_normal_log("write done!")
        except Exception as e:
            self.textBrowser_error_log('write err:%s' % e)

    def write_atom(self, addr, data):
        try:
            write_buffer = (c_ubyte * 8)()
            addr_str = '{:0>8x}'.format(addr * 2048)
            data_str = '{:0>8x}'.format(data)
            write_buffer[0] = int(addr_str[0:2], 16)
            write_buffer[1] = int(addr_str[2:4], 16)
            write_buffer[2] = int(addr_str[4:6], 16)
            write_buffer[3] = int(addr_str[6:], 16)
            write_buffer[4] = int(data_str[0:2], 16)
            write_buffer[5] = int(data_str[2:4], 16)
            write_buffer[6] = int(data_str[4:6], 16)
            write_buffer[7] = int(data_str[6:8], 16)
            nRet = ControlSPI.VSI_WriteBytes(ControlSPI.VSI_USBSPI, 0, 0, write_buffer, 8)
            # self.textBrowser_normal_log('write done!')
        except Exception as e:
            self.textBrowser_error_log('write_atom err:%s' % e)

    def read_atom(self, addr):
        try:
            write_buffer = (c_ubyte * 4)()
            read_value = (c_ubyte * 4)()
            addr_str = '{:0>8x}'.format(addr * 2048)
            write_buffer[0] = int(addr_str[0:2], 16) + 128
            write_buffer[1] = int(addr_str[2:4], 16)
            write_buffer[2] = int(addr_str[4:6], 16)
            write_buffer[3] = int(addr_str[6:8], 16)
            nRet = ControlSPI.VSI_WriteReadBytes(ControlSPI.VSI_USBSPI, 0, 0, write_buffer, 4, read_value, 4)
            # self.textBrowser_normal_log("read done!")
            return read_value
        except Exception as e:
            self.textBrowser_error_log('read_atom err:%s' % e)

    def load_test_seq(self):
        # try:
        self.textBrowser_normal_log("导入一张表单信息")
        test_seq_file, filetype = QFileDialog.getOpenFileName(self, "choose file", "./",
                                                              "All Files (*);;excel Files (*.xlsx);;excel Files ("
                                                              "*.xls)")  # 设置文件扩展名过滤,注意用双分号间隔
        if test_seq_file == "":
            return
        else:
            self.parser_seq_file(test_seq_file)
            self.textBrowser_normal_log("open a file:%s" % test_seq_file)
        # except Exception as e:
        #     self.textBrowser_error_log("parser_seq_file err:%s" % e)

    def deal_emit_sheet(self, select_sheet):
        self.sheet_sel_lst = select_sheet

    def just_open(self, filename):
        xlApp = win32com.client.DispatchEx('Excel.Application')
        xlApp.Visible = False  # 是否可视化编辑
        xlApp.ScreenUpdating = False  # 画面刷新显示
        xlApp.DisplayAlerts = False
        if os.path.exists(filename):
            xlBook = xlApp.Workbooks.Open(filename, False)
            xlBook.Save()
            xlBook.Close()
            xlApp.Quit()

    def parser_seq_file(self, fn):
        self.just_open(fn)
        data = op.load_workbook(fn, data_only=True)
        sheetsall = data.get_sheet_names()
        dialog = MainDialog1(sheetsall)
        dialog.Signal_parp.connect(self.deal_emit_sheet)
        dialog.show()
        dialog.exec_()
        sheet_idx = sheetsall.index(self.sheet_sel_lst)
        sheet_data = data.get_sheet_by_name(sheetsall[sheet_idx])
        rows_num = sheet_data.max_row
        for x in range(1, rows_num + 1):
            if sheet_data.cell(x, 1).value == 'sleep':
                if sheet_data.cell(x, 2).value == '':
                    time.sleep(5)
                else:
                    tmp_time = int(sheet_data.cell(x, 2).value)
                    time.sleep(tmp_time)
            elif sheet_data.cell(x, 1).value == 'wait':
                # temp_addr = int(sheet_data.cell(x, 2).value, 16)+self.base_addr_dict[self.choose_path][self.dac_tile_comboBox.currentIndex()]
                temp_addr = int(sheet_data.cell(x, 2).value, 16)
                temp_value = int(sheet_data.cell(x, 3).value, 16)
                for i in range(0, 300):
                    time.sleep(1)
                    read_value = self.read_atom(temp_addr)
                    if read_value[0] * 16777216 + read_value[1] * 65536 + read_value[2] * 256 + read_value[
                        3] == temp_value:
                        break
            elif sheet_data.cell(x, 1).value is None:
                pass
            else:
                # temp_addr = int(sheet_data.cell(x, 1).value, 16) + self.base_addr_dict[self.choose_path][self.dac_tile_comboBox.currentIndex()]
                temp_addr = int(sheet_data.cell(x, 1).value, 16)
                temp_data = int(sheet_data.cell(x, 2).value, 16)
                self.write_atom(temp_addr, temp_data)
                time.sleep(0.01)

    def parser_no_choose_seq_file(self, fn, sheet_name, base_addr):
        # self.just_open(fn)
        data = op.load_workbook(fn, data_only=True)
        sheet_data = data.get_sheet_by_name(sheet_name)
        rows_num = sheet_data.max_row
        for x in range(1, rows_num + 1):
            if sheet_data.cell(x, 1).value == 'sleep':
                if sheet_data.cell(x, 2).value == '':
                    time.sleep(5)
                else:
                    tmp_time = int(sheet_data.cell(x, 2).value)
                    time.sleep(tmp_time)
            elif sheet_data.cell(x, 1).value == 'wait':
                temp_addr = int(sheet_data.cell(x, 2).value, 16)+self.base_addr_dict[self.choose_path][self.dac_tile_comboBox.currentIndex()]
                # temp_addr = int(sheet_data.cell(x, 2).value, 16) + base_addr
                temp_value = int(sheet_data.cell(x, 3).value, 16)
                for i in range(0, 300):
                    time.sleep(1)
                    read_value = self.read_atom(temp_addr)
                    if read_value[0] * 16777216 + read_value[1] * 65536 + read_value[2] * 256 + read_value[
                        3] == temp_value:
                        break
            elif sheet_data.cell(x, 1).value is None:
                pass
            else:
                temp_addr = int(sheet_data.cell(x, 1).value, 16) + self.base_addr_dict[self.choose_path][self.dac_tile_comboBox.currentIndex()]
                # temp_addr = int(sheet_data.cell(x, 1).value, 16) + base_addr
                temp_data = int(sheet_data.cell(x, 2).value, 16)
                self.write_atom(temp_addr, temp_data)
                time.sleep(0.01)

    def read_mem_atom(self, addr):
        write_buffer = (c_ubyte * 4)()
        read_value = (c_ubyte * 4)()
        addr_str = '{:0>8x}'.format(addr * 2048)
        write_buffer[0] = int(addr_str[0:2], 16) + 128
        write_buffer[1] = int(addr_str[2:4], 16)
        write_buffer[2] = int(addr_str[4:6], 16)
        write_buffer[3] = int(addr_str[6:8], 16)
        nRet = ControlSPI.VSI_WriteReadBytes(ControlSPI.VSI_USBSPI, 0, 0, write_buffer, 4, read_value, 4)
        return read_value

    def wirte_to_txt_sample(self):
        self.textBrowser_normal_log("数采开始!")
        base_addr = self.base_addr_dict[self.choose_path][self.adc_pll_tile_comboBox.currentIndex()]
        ###################################################################################################
        path_DAC_reg = r'DATA_SAMPLE_XIHEV200.xlsx'
        wb_reg_workbook = op.load_workbook(resource_path(os.path.join("source", path_DAC_reg)))
        table_adc_setting = wb_reg_workbook.get_sheet_by_name('XIHEV200_ADC_MEMORYREAD')
        if self.adc_chn_choose_comboBox.currentIndex() == 0:
            table_adc_setting.cell(16, 5, 1)
            table_adc_setting.cell(19, 5, 0)
            table_adc_setting.cell(21, 5, 0)
            table_adc_setting.cell(23, 5, 0)
        elif self.adc_chn_choose_comboBox.currentIndex() == 1:
            table_adc_setting.cell(16, 5, 0)
            table_adc_setting.cell(19, 5, 1)
            table_adc_setting.cell(21, 5, 0)
            table_adc_setting.cell(23, 5, 0)
        elif self.adc_chn_choose_comboBox.currentIndex() == 2:
            table_adc_setting.cell(16, 5, 0)
            table_adc_setting.cell(19, 5, 0)
            table_adc_setting.cell(21, 5, 1)
            table_adc_setting.cell(23, 5, 0)
        else:
            table_adc_setting.cell(16, 5, 0)
            table_adc_setting.cell(19, 5, 0)
            table_adc_setting.cell(21, 5, 0)
            table_adc_setting.cell(23, 5, 1)
        wb_reg_workbook.save(resource_path(os.path.join("source", path_DAC_reg)))
        wb_reg_workbook.close()
        self.just_open(resource_path(os.path.join("source", path_DAC_reg)))

        path_DAC_reg = r'DATA_SAMPLE_XIHEV200.xlsx'
        wb_reg = op.load_workbook(resource_path(os.path.join("source", path_DAC_reg)),data_only=True)
        reg_sheet = wb_reg.get_sheet_by_name('XIHEV200_SMP_INIT')
        self.write_atom(base_addr + 0x40b, 0xf)  #
        self.write_atom(base_addr + 0x4e1, 0x0)  #
        # cols = 2
        rows = reg_sheet.max_row
        addr_buffer = []
        data_buffer = []
        for row in range(1,rows+1):
            if reg_sheet.cell(row, 1).value is not None:
                addr_buffer.append(int(reg_sheet.cell(row, 1).value, 16))
                data_buffer.append(int(reg_sheet.cell(row, 2).value, 16))
        # print(addr_buffer, data_buffer)
        # wb_reg.close()
        # spi = op_spi(freq=1125000)
        # self.spi_config()
        for i in range(0, 29):
            # self.read_atom(base_addr+addr_buffer[i])
            self.write_atom(base_addr + addr_buffer[i], data_buffer[i])
            time.sleep(0.01)
            # self.read_atom(base_addr+0x1)
            self.read_atom(base_addr + addr_buffer[i])
            time.sleep(0.01)

        # path_DAC_reg = r'DATA_SAMPLE_XIHEV200.xlsx'
        # wb_reg = op.load_workbook(resource_path(os.path.join("source", path_DAC_reg)),data_only=True)
        reg_sheet = wb_reg.get_sheet_by_name('XIHEV200_ADC_MEMORYREAD')

        # cols = 2
        rows = reg_sheet.max_row
        addr_buffer = []
        data_buffer = []

        for row in range(1, rows+1):
            if reg_sheet.cell(row, 1).value is not None:
                addr_buffer.append(int(reg_sheet.cell(row, 1).value, 16))
                data_buffer.append(int(reg_sheet.cell(row, 2).value, 16))
        # print(addr_buffer, data_buffer)
        wb_reg.close()
        # spi = op_spi(freq=1125000)
        # self.spi_config()
        for i in range(len(addr_buffer)):
            # self.read_atom(base_addr+addr_buffer[i])
            self.write_atom(base_addr + addr_buffer[i], data_buffer[i])
            time.sleep(0.01)
            # self.read_atom(base_addr+0x1)
            self.read_atom(base_addr + addr_buffer[i])
            time.sleep(0.01)

        #################################################### data sample ###################################

        # self.write_atom(base_addr + 0x4c8, 0x0)
        # self.write_atom(base_addr + 0x4c4, 0x0)
        # self.write_atom(base_addr + 0x4c1, 0x2)
        # self.write_atom(base_addr+0x470, 0x0)
        # self.write_atom(base_addr+0x470, 0x40000)
        # self.write_atom(base_addr+0x470, 0x0)
        read_len = 4  # byte
        memory_wid = 4  # byte, 32bit
        memory_depth = 32 * 1024 * memory_wid  # by0te
        addr_step = int(read_len / memory_wid)
        read_times = int(memory_depth / read_len)
        fp = open('memory_dump_data.txt', 'w')
        return_value = (c_ubyte * 4)()
        for i in range(0, read_times):
            self.write_atom(base_addr + 0x4c0, i * 4 * addr_step)
            self.write_atom(base_addr + 0x4c0, 0x40000 + i * 4 * addr_step)
            self.write_atom(base_addr + 0x4c0, i * 4 * addr_step)
            self.write_atom(base_addr + 0x4c0, i * 4 * addr_step)
            self.write_atom(base_addr + 0x4c0, 0x40000 + i * 4 * addr_step)
            self.write_atom(base_addr + 0x4c0, i * 4 * addr_step)
            read_buffer = self.read_mem_atom(base_addr + 0x4cC)
            for k in range(0, int(len(read_buffer) / 4)):
                fp.write("0x")
                for j in range(0, 4):
                    fp.write("%02x" % (read_buffer[k * 4 + j]))
                fp.write("\n")
        fp.close()
        self.write_atom(base_addr + 0x4c1, 0x0)

        ########################## read calibration #########################

        self.read_atom(base_addr + 0x1068D)
        self.read_atom(base_addr + 0x1068E)
        self.read_atom(base_addr + 0x1068F)
        self.read_atom(base_addr + 0x10690)
        self.read_atom(base_addr + 0x10691)
        self.read_atom(base_addr + 0x10692)
        self.read_atom(base_addr + 0x10693)
        self.read_atom(base_addr + 0x10694)

        #############
        MEM_SEC_NUM = 32
        MEM_SEC_SIZE = 2048

        ##need config
        start_sec_idx = 0
        end_sec_idx = 16
        bit_num = 2

        read_file_name = 'memory_dump_data.txt'
        write_file_name = 'read_data_from_testmem_after_trans.txt'

        smp_width = (2 ** bit_num) * 32
        smp_sec_num = 2 ** bit_num
        sec_use_num = end_sec_idx - start_sec_idx + 1
        sec_grp_num = int(sec_use_num / smp_sec_num)

        with open(read_file_name, "r") as fn:
            rdata_buf = fn.readlines()
        for i in range(0, len(rdata_buf)):
            rdata_buf[i] = int(rdata_buf[i].rstrip('\n'), 16)

        wdata_buf = []

        cnt = 0
        self.textBrowser_normal_log('sec_grp_num=' + str(sec_grp_num) + '\n')
        self.textBrowser_normal_log('MEM_SEC_SIZE=' + str(MEM_SEC_SIZE) + '\n')
        self.textBrowser_normal_log('smp_sec_num=' + str(smp_sec_num) + '\n')
        for sec_grp_idx in range(0, sec_grp_num):
            for sec_depth_idx in range(0, MEM_SEC_SIZE):
                for smp_sec_idx in range(0, smp_sec_num):
                    wdata_buf.append(rdata_buf[
                                         sec_grp_idx * smp_sec_num * MEM_SEC_SIZE + smp_sec_idx * MEM_SEC_SIZE + sec_depth_idx])
                    # fdbg.write('cnt='+str(cnt)+',  rdata_idx='+str(sec_grp_idx*smp_sec_num*MEM_SEC_SIZE + smp_sec_idx*MEM_SEC_SIZE + sec_depth_idx)+',  rdata='+str(wdata_buf[cnt])+'\n')
                    cnt = cnt + 1

        fp = open(write_file_name, 'w')
        wdata = []
        for i in range(0, len(wdata_buf)):
            fp.write("%02x\n" % (wdata_buf[i] % 65536))
            fp.write("%02x\n" % ((int(wdata_buf[i] / 65536)) % 65536))
            wdata.append(wdata_buf[i] % 65536)
            wdata.append((int(wdata_buf[i] / 65536)) % 65536)
        fp.close()
        self.textBrowser_normal_log("数采finish!")
